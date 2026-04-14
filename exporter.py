"""
LeadHarvester Pro - Excel Exporter
====================================
Export scraped leads to professionally formatted Excel files
with styling, conditional formatting, and summary sheets.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from loguru import logger

from config import OUTPUT_DIR


class LeadExporter:
    """
    Export leads to beautifully formatted Excel & CSV files.

    Usage:
        exporter = LeadExporter()
        path = exporter.to_excel(leads, keyword="restaurants", location="NYC")
        path = exporter.to_csv(leads, keyword="restaurants", location="NYC")
    """

    # Column display configuration
    COLUMN_CONFIG = {
        "name": {"label": "Business Name", "width": 30},
        "industry_category": {"label": "Industry", "width": 25},
        "address": {"label": "Address", "width": 40},
        "phone": {"label": "Phone", "width": 18},
        "website": {"label": "Website", "width": 35},
        "rating": {"label": "Rating", "width": 10},
        "review_count": {"label": "Reviews", "width": 12},
        "lead_quality_score": {"label": "Lead Score", "width": 12},
        "price_level": {"label": "Price Level", "width": 12},
        "data_confidence": {"label": "Confidence", "width": 13},
        "estimated_revenue": {"label": "Est. Revenue", "width": 15},
        "outreach_angle": {"label": "Outreach Hook", "width": 50},
        "google_maps_url": {"label": "Google Maps Link", "width": 40},
        "scraped_at": {"label": "Scraped At", "width": 20},
    }

    def to_excel(
        self,
        leads: list[dict],
        keyword: str = "",
        location: str = "",
        output_dir: Path | None = None,
        filename: str | None = None,
    ) -> Path:
        """Export leads to a styled Excel workbook."""
        out_dir = output_dir or OUTPUT_DIR
        out_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_kw = keyword.replace(" ", "_")[:20] if keyword else "leads"
        safe_loc = location.replace(" ", "_").replace(",", "")[:15] if location else ""
        fname = filename or f"{safe_kw}_{safe_loc}_{timestamp}"
        filepath = out_dir / f"{fname}.xlsx"

        # Build DataFrame with ordered columns
        df = pd.DataFrame(leads)
        ordered_cols = [c for c in self.COLUMN_CONFIG if c in df.columns]
        # Add any extra columns not in config
        extra_cols = [c for c in df.columns if c not in self.COLUMN_CONFIG]
        all_cols = ordered_cols + extra_cols
        df = df[all_cols]

        # Create workbook
        wb = Workbook()

        # Sheet 1: All Leads
        ws_all = wb.active
        ws_all.title = "All Leads"
        self._write_styled_sheet(ws_all, df, leads)

        # Sheet 2: High Quality Leads (score >= 7)
        if "lead_quality_score" in df.columns:
            high_quality = df[df["lead_quality_score"] >= 7].copy()
            if not high_quality.empty:
                ws_hq = wb.create_sheet("Hot Leads")
                self._write_styled_sheet(ws_hq, high_quality, high_quality.to_dict("records"))

        # Sheet 3: Summary Dashboard
        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary, df, keyword, location)

        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath} ({len(leads)} leads)")
        return filepath

    def to_csv(
        self,
        leads: list[dict],
        keyword: str = "",
        location: str = "",
        output_dir: Path | None = None,
    ) -> Path:
        """Export leads to CSV."""
        out_dir = output_dir or OUTPUT_DIR
        out_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_kw = keyword.replace(" ", "_")[:20] if keyword else "leads"
        filepath = out_dir / f"{safe_kw}_{timestamp}.csv"

        df = pd.DataFrame(leads)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")  # BOM for Excel compat

        logger.info(f"CSV saved: {filepath} ({len(leads)} leads)")
        return filepath

    # ── Private: Styled Sheet ────────────────────────────────

    def _write_styled_sheet(
        self, ws, df: pd.DataFrame, leads: list[dict]
    ) -> None:
        """Write DataFrame to worksheet with professional styling."""
        # Colors
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        score_high = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        score_med = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        score_low = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        num_cols = len(df.columns)
        num_rows = len(df) + 1

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            label = self.COLUMN_CONFIG.get(col_name, {}).get("label", col_name.replace("_", " ").title())
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if pd.isna(value) or value is None:
                    cell.value = ""
                elif isinstance(value, bool):
                    cell.value = "Yes" if value else "No"
                else:
                    cell.value = value

                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                # Alternating rows
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

                # Color-code lead quality score
                if col_name == "lead_quality_score" and isinstance(value, (int, float)):
                    if value >= 7:
                        cell.fill = score_high
                    elif value >= 4:
                        cell.fill = score_med
                    else:
                        cell.fill = score_low
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Center ratings
                if col_name in ("rating", "review_count", "data_confidence"):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            width = self.COLUMN_CONFIG.get(col_name, {}).get("width", 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row + first column
        ws.freeze_panes = "B2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"

        # Row height
        ws.row_dimensions[1].height = 25

    def _write_summary_sheet(
        self, ws, df: pd.DataFrame, keyword: str, location: str
    ) -> None:
        """Write a summary dashboard sheet."""
        title_font = Font(name="Calibri", bold=True, size=14, color="0F172A")
        label_font = Font(name="Calibri", bold=True, size=11, color="334155")
        value_font = Font(name="Calibri", size=11, color="0F172A")
        accent_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value="LeadHarvester Pro - Scrape Summary")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center")

        # Metadata
        data = [
            ("Search Query", keyword),
            ("Location", location),
            ("Date Scraped", datetime.now().strftime("%B %d, %Y %I:%M %p")),
            ("Total Leads", len(df)),
        ]

        # Stats
        if "rating" in df.columns:
            avg_rating = df["rating"].dropna().mean()
            if pd.notna(avg_rating):
                data.append(("Avg Rating", f"{avg_rating:.1f}"))

        if "lead_quality_score" in df.columns:
            hot = len(df[df["lead_quality_score"] >= 7])
            data.append(("Hot Leads (Score 7+)", hot))
            data.append(("Avg Lead Score", f"{df['lead_quality_score'].mean():.1f}"))

        if "phone" in df.columns:
            has_phone = df["phone"].notna().sum() - (df["phone"] == "").sum()
            data.append(("With Phone Number", int(has_phone)))

        if "website" in df.columns:
            has_web = df["website"].notna().sum() - (df["website"] == "").sum()
            data.append(("With Website", int(has_web)))

        if "industry_category" in df.columns:
            top_cats = df["industry_category"].value_counts().head(5)
            data.append(("", ""))
            data.append(("Top Categories", ""))
            for cat, count in top_cats.items():
                data.append((f"  {cat}", count))

        for row_idx, (label, value) in enumerate(data, 3):
            ws.cell(row=row_idx, column=1, value=label).font = label_font
            ws.cell(row=row_idx, column=2, value=value).font = value_font
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=1).fill = accent_fill
                ws.cell(row=row_idx, column=2).fill = accent_fill

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 35
